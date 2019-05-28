from tkinter import *
from tkinter import messagebox,filedialog,ttk
import os, sys, openpyxl,time,pygame
from PIL import ImageTk, Image
import tkinter as tk
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
import openpyxl
import matplotlib.pylab as plt
pygame.init()
mypath = os.path.dirname(os.path.realpath(__file__))
click1= pygame.mixer.Sound(mypath+"\\Sound\\223.wav")
click2= pygame.mixer.Sound(mypath+"\\Sound\\213.wav")
def center(win):
    win.update_idletasks()
    width = win.winfo_width()
    frm_width = win.winfo_rootx() - win.winfo_x()
    win_width = width + 2 * frm_width
    height = win.winfo_height()
    titlebar_height = win.winfo_rooty() - win.winfo_y()
    win_height = height + titlebar_height + frm_width
    x = win.winfo_screenwidth() // 2 - win_width // 2
    y = win.winfo_screenheight() // 2 - win_height // 2
    win.geometry('{}x{}+{}+{}'.format(width, height, x, y))
    win.deiconify()
    
def register():
    global  ID,email,password2,username,register_screen,var, var1,var2,c,D,droplist2,list2,label_6,submit,phone, list1
    main_screen.destroy()
    
#Regestration Screen
    register_screen = Tk()
    register_screen.title("Register")
    register_screen.geometry("861x678")
    register_screen.grid_columnconfigure(1,weight=1)
    register_screen.resizable(0, 0)
    
    pygame.mixer.Sound(mypath+"\\Sound\\Pleasefill.wav").play()
    
    
# Regestration Header
    cv2 = tk.Canvas(register_screen, width=2000, height=2000)
    cv2.pack(fill='both')
    img_file2 = mypath+"\\Images\\Registration.jpg"
    pil_image2 = Image.open(img_file2)  
    tk_image2 = ImageTk.PhotoImage(pil_image2)
    cv2.create_image(430, 343, image=tk_image2, anchor='center') 

#ID
    ID = StringVar()
    ID = Entry(register_screen,width=37,borderwidth=0)
    ID.place(x=510,y=224)
    
#Name
    username = StringVar()
    username = Entry(register_screen,width=37,borderwidth=0)
    username.place(x=150,y=224)
     
#Password
    password2 = StringVar()
    password2 = Entry(register_screen, show= '*',width=36,borderwidth=0)
    password2.place(x=510,y=322)
    
#Email
    email=StringVar()
    email = Entry(register_screen,width=36,borderwidth=0)
    email.place(x=150,y=322)
#Phone
    phone = StringVar() 
    phone = Entry(register_screen,width=36,borderwidth=0)
    phone.place(x=150,y=416)   
#Gender
    var = IntVar()
    Radiobutton(register_screen, text="Male",padx = 5, variable=var, value=1 ,background="white",activebackground="white",width=15).place(x=490,y=413)
    Radiobutton(register_screen, text="Female",padx = 20, variable=var, value=2,background="white",activebackground="white").place(x=590,y=413)
#School droplist
    c=StringVar()
    list1 = ['Business Administration (BA)','Information Technology and CS (ITCS)','Engineering and Applied Sciences (EAS)']
    droplist=OptionMenu(register_screen,c, *list1)
    droplist.config(width=37,activebackground="white",bg="white",highlightthickness=0)
    droplist ["menu"] ["bg"] = "white"
    c.set('select your School') 
    droplist.place(x=112,y=511)
    c.trace('w', option_select)
#Major droplist
    list2=["Select School First"]
    label_6 = Label(register_screen, text="Bachelor",width=20,font=("bold", 13),bg="white")
    D=StringVar()
    droplist2=OptionMenu(register_screen,D, *list2)
    droplist2.config(width=40,activebackground="white",bg="white",highlightthickness=0)
    droplist2 ["menu"] ["bg"] = "white"
    D.set('select your Major')        
    droplist2.place(x=470,y=511)
    label_6.place(x=410,y=477)
    
    submit=Button(register_screen,font="Myriad 10 bold", text='Submit',width=28,height=1,bg='#002145',fg='white',command=check_reg_data)
    submit.place(x=315,y=585)
    
    
    label_7 = Label(register_screen, text="Back",width=8,font=("bold", 13),bg="#002145",fg="white")
    label_7.bind("<Button-1>", back)
    label_7.place(x=27,y=37)
    center(register_screen)    
    register_screen.mainloop()    
    
def check_reg_data() :
    if ID.get() =="" or username.get()=="" or password2.get()=="" or email.get()=="" or phone.get()=="" or var.get()== False :
        messagebox.showwarning("Error", "Please fill all required fields")
    elif c.get() not in list1:
        messagebox.showwarning("Error", "Please fill all required fields")
    elif c.get() =='Information Technology and CS (ITCS)' and D.get() not in listITCS:
        messagebox.showwarning("Error", "Please fill all required fields")
    elif c.get() =='Engineering and Applied Sciences (EAS)' and D.get() not in listEAS:
        messagebox.showwarning("Error", "Please fill all required fields")
    else:
        registration_sucess()

def option_select(*args):
    global list2,droplist2,list2,label_6,listEAS,listITCS
#Major lists
    listBA = [];
    listITCS = ['Science in Computer Science (CS)','Science in Biomedical Informatics'];
    listEAS = ['Science in Industrial Engineering and Management (ISEM)','Science in Mechanical Engineering (MENG)','Science in Civil & Infrastructure Engineering & Management (CIEM)','Science in Electronics and Computer Engineering','Science in Architecture Engineering and Urban Design']
      
#Don't even try to understand
    try:
        label_6
    except NameError:
        label_6 = None
        
    if label_6 == None:
        
        
        if c.get()== 'Information Technology and CS (ITCS)':
            label_6 = Label(register_screen, text="Bachelor",width=20,font=("bold", 13),bg="white")
            list2=["Select School First"]
            D=StringVar()
            droplist2=OptionMenu(register_screen,D, *list2)
            droplist2.config(width=40,activebackground="white",bg="white",highlightthickness=0)
            droplist2 ["menu"] ["bg"] = "white"
            D.set('select your Major')        
            label_6.place(x=410,y=477)
            droplist2.place(x=470,y=511)   
            list2=[]   
            for i in listITCS:
                list2.append(i)
            update_option_menu()
        elif c.get()== 'Engineering and Applied Sciences (EAS)':
            label_6 = Label(register_screen, text="Bachelor",width=20,font=("bold", 13),bg="white")
            list2=["Select School First"]
            D=StringVar()
            droplist2=OptionMenu(register_screen,D, *list2)
            droplist2.config(width=40,activebackground="white",bg="white",highlightthickness=0)
            droplist2 ["menu"] ["bg"] = "white"
            D.set('select your Major')        
            label_6.place(x=410,y=477)
            droplist2.place(x=470,y=511)       
            list2=[]   
            for i in listEAS:
                list2.append(i)
            update_option_menu()
        
    else:
        
        if c.get()== 'Business Administration (BA)':
            label_6.destroy()
            droplist2.destroy()
            label_6= None
        elif c.get()== 'Information Technology and CS (ITCS)':
                   
            list2=[]   
            for i in listITCS:
                list2.append(i)
            update_option_menu()
        elif c.get()== 'Engineering and Applied Sciences (EAS)':
                  
            list2=[]   
            for i in listEAS:
                list2.append(i)
            update_option_menu() 
        

def back(Event):
    click2.play()
    register_screen.destroy()
    main_account_screen()
def update_option_menu():
    menu = droplist2["menu"]
    menu.delete(0, "end")
    for i in list2:
        menu.add_command(label=i, command=lambda value=i:D.set(value))    
 

 
def login():
    #Label1.destroy()
    Label2.destroy()
    NewButton.destroy()
    NewButton2.destroy()
    
    global username_verify
    global password_verify

    username_verify = StringVar()
    password_verify = StringVar()

    global username_login_entry
    global password_login_entry
    Label(main_screen, text="Please enter details below to login",bg="white",highlightthickness=0).place(x=155,y=234)
    Label(main_screen, text="Username * ",bg="white",highlightthickness=0).place(x=210,y=263)
    username_login_entry = Entry(main_screen, textvariable=username_verify,bg="white",highlightthickness=0)
    username_login_entry.place(x=180,y=290)
    Label(main_screen, text="Password * ",bg="white",highlightthickness=0).place(x=210,y=316)
    password_login_entry = Entry(main_screen, textvariable=password_verify, show= '*',bg="white",highlightthickness=0)
    password_login_entry.place(x=180,y=343)
    bot=Button(main_screen, text="Login", width=11, height=1,bg="white",highlightthickness=0)
    bot.place(x=197,y=380)
    bot.bind("<Button-1>",login_verify)
    main_screen.bind("<Return>",login_verify)
    Label(main_screen, text="Don't have an account yet?",bg="white",highlightthickness=0).place(x=135,y=417)
    l5=Label(main_screen, text="Register..", fg="blue",bg="white",highlightthickness=0)
    l5.place(x=285,y=417)
    l5.bind("<Button-1>", openreg)
 
def openreg(Event):
    register()
    
def login_verify(Event):
    global username1,s
    username1 = username_verify.get()
    password1 = password_verify.get()
    from pathlib import Path
    path=os.getcwd()
    list_of_files = os.listdir(mypath+"\\Users")
    l=[]
    for i in list_of_files:
        x=i.find(".")
        l=l+[i[ :x]]
        
    if username1 in l:
        file1 = open(mypath+"\\Users\\"+username1+".txt", "r")
        verify = file1.readlines()
        s=[]
        for i in range (len(verify)):
            s.append(verify[i].strip())
        if "Password: "+password1 in s:
            username_login_entry.delete(0, END)
            password_login_entry.delete(0, END)            
            messagebox.showinfo("Success", "Login Success\n welcome, %s"%(username1.split(" ")[0]))
            main_screen.destroy()
            Home_screenf()            
        else:
            password_not_recognised()
            password_login_entry.delete(0, END)            
 
    else:
        messagebox.showwarning("Error", "User Not Found")
        username_login_entry.delete(0, END)
        password_login_entry.delete(0, END)        
 
   
 
def registration_sucess():
    username_info = username.get()
    password_info = password2.get()
    Email= email.get()
    if var.get() == 1:
        gender="male"
    else:
        gender="female"
        
#Writing
    file = open(mypath+"\\Users\\"+username_info+".txt", "w")
    file.write("Name: "+username.get())
    file.write("\nID: "+ID.get())
    file.write("\nEmail: "+email.get())
    file.write("\nPassword: "+password_info)
    file.write("\nGender: "+ gender)
    file.write("\nSchool of "+ c.get())
    if D.get()!="select your Major":
        file.write("\nBachelor of "+ D.get())
   
    file.close()
    messagebox.showinfo("Success", "Registration Success")
    register_screen.destroy()
    main_account_screen()    
    
    
#---------------------------------------------------------------------------------------------------------------------------------
    

def back3(Event):
    financial_screen.destroy()
    v1,v5,v6,v8,v9,v10,v11,v13,v14,v15,v16,v17,v18=N1,N5,N6,N8,N9,M1,M2,M4,Q9,K2,K3,K4,grade
    financialscreen(v1,v5,v6,v8,v9,v10,v11,v13,v14,v15,v16,v17,v18)

    
def afterprogress():
    progress.stop()
    progress.destroy()
    tex2.destroy()
    labelback3 = Label(financial_screen, text="Back",width=8,font=("bold", 13),bg="#002145",fg="white")
    labelback3.bind("<Button-1>", back2)
    labelback3.place(x=424,y=10)
    cv2.itemconfig(background, image = defaultimg3)
    pygame.mixer.Sound(mypath+"\\Sound\\analysiscomp.wav").play()   
      
    tex3=Label(financial_screen,text=stuname, width="20", height="2", font=("Lato", 14),background="white",activebackground="white")
    #tex3.place(x=352,y=251)
    tex3.place(x=363,y=199)
    tex3.config(anchor='w')
    
    tex4=Label(financial_screen,text=stuID, width="20", height="2", font=("Lato", 14),background="white",activebackground="white")
    tex4.place(x=363,y=244)
    tex4.config(anchor='w')
    for i in s:
        x=i.split(" ")
        if "School" in x:
            c=x[2: ]
            stuschool=" ".join(c)    
    tex5=Label(financial_screen,text=stuschool, width="41", height="2", font=("Lato", 14),background="white",activebackground="white")
    tex5.place(x=363,y=289)
    tex5.config(anchor='w')
    if stuschool!= "Business Administration (BA)":
        for i in s:
            x=i.split(" ")
            if "Bachelor" in x:
                c=x[2: ]
                stubat=" ".join(c)
                
    else:
        stubat="No options"
    tex6=Label(financial_screen,text=stubat, width="39", height="2", font=("Lato", 14),background="white",activebackground="white")
    tex6.place(x=363,y=334)
    tex6.config(anchor='w')   
    tex7=Label(financial_screen,text=str(grade)+" %", width="10", height="2", font=("Lato", 14),background="white",activebackground="white")
    tex7.place(x=363,y=380)
    tex7.config(anchor='w')
    tex8=Label(financial_screen,text=str(TotalN)+" EGP", width="15", height="2", font=("Lato", 14),background="white",activebackground="white")
    tex8.place(x=363,y=428)
    tex8.config(anchor='w')
    tex9=Label(financial_screen,text=str(TotalD)+" EGP", width="15", height="2", font=("Lato", 14),background="white",activebackground="white")
    tex9.place(x=363,y=473)
    tex9.config(anchor='w')
    tex10=Label(financial_screen,text=str(TotalM)+" EGP", width="15", height="2", font=("Lato", 14),background="white",activebackground="white")
    tex10.place(x=363,y=518)
    tex10.config(anchor='w')
    tex11=Label(financial_screen,text=str(S1)+" %", width="4", height="1", font=("Lato", 14),background="white",activebackground="white")
    tex11.place(x=590,y=606)
    tex11.config(anchor='w')
    tex12=Label(financial_screen,text=str(S2)+" %", width="4", height="1", font=("Lato", 14),background="white",activebackground="white")
    tex12.place(x=590,y=643)
    tex12.config(anchor='w')    
def changetext():
    global tex2
    tex1.destroy()
    tex2=Label(text="Analysing financial information ", width="30", height="2", font=("Lato", 12),background="white",activebackground="white")
    tex2.place(x=330,y=280)    
def results():
    global progress,tex1
    e1.destroy()
    if e2.get()==1:
        e22.destroy()
    if e3.get()==1:
        e33.destroy()
    if e4.get()==1:
        e44.destroy()
    e5.destroy()
    e6.destroy()
    e8.destroy()
    e9.destroy()
    e10.destroy()
    e11.destroy()
    if e12.get()==1 or e12.get()==2 :
        e1212.destroy()
    e13.destroy()
    e14.destroy()
    e15.destroy()
    e16.destroy()
    e17.destroy()
    e18.destroy()
    r1.destroy()
    r2.destroy()
    r3.destroy()
    r4.destroy()
    r5.destroy()
    r6.destroy()
    r7.destroy()
    r8.destroy()
    submit.destroy()
    labelback2.destroy()
    studentnamelabel.destroy()
    studentname.destroy()
    studentidlabel.destroy()
    studentnamelabel.destroy()
    studentid.destroy()
    cv2.itemconfig(background, image = defaultimg2)
    

    progress = ttk.Progressbar(financial_screen, orient = HORIZONTAL, length = 450)
    progress.place(x=236,y=254)
    progress.config(mode = 'determinate',maximum=70, value = 1)
    progress.step(1)
    pygame.mixer.Sound(mypath+"\\Sound\\analysing.wav").play() 
    progress.start()
    tex1=Label(text="Collecting student records", width="20", height="2", font=("Lato", 12),background="white",activebackground="white")
    tex1.place(x=375,y=280)
    progress.after(3400, changetext)
    progress.after(6900, afterprogress)    
    
def getdata(Event):
    global TotalN,TotalM,TotlaK,N1,N5,N6,N8,N9,M1,M2,M4,Q9,K2,K3,K4,grade,S1,S2,TotalD,TotalK,TotalM,TotalN,Total_scholarship
    if e1.get() == "" or e2.get() == False or e2.get() == "" or e3.get() == False or e3.get() == "" or e4.get() == False or e4.get() == "" or e5.get() == "" or e6.get() == "" or e8.get() == "" or e9.get() == "" or e10.get() == "" or e11.get() == "" or e12.get() == False or e12.get() == "" or e13.get() == "" or e14.get() == "" or e15.get() == "" or e16.get() == "" or e17.get() == "" or e18.get()==""  :
        messagebox.showwarning("Error", "Please fill all required fields")
    elif float(e18.get())>100:
        messagebox.showwarning("Error", "Please enter valid grade")
    else:
        click1.play()
        N1=e1.get()
        N5=e5.get()
        N6=e6.get()
        N8=e8.get()
        N9=e9.get()
        M1=e10.get()
        M2=e11.get()
        M4=e13.get()
        Q9=e14.get()
        K2=e15.get()
        K3=e16.get()
        K4=e17.get()
        grade=float(e18.get())
        if e2.get() ==1 :
            N2=e22.get()
        elif e3.get() ==2:
            N2=0
        if e3.get() ==1 :
            N3=e33.get()
        elif e3.get() ==2:
            N3=0    
        
        if e4.get() ==1 :
            N4=e44.get()     
        elif e4.get() ==2:
            N4=0
    
        if e12.get() ==1 :
            M3=0
            K1=e1212.get()        
        elif e12.get() ==2:
            M3=e1212.get()  
            K1=0
  
    #Total Annual dynamic resources
        TotalN= float(N1) + float(N2) + float(N3) + float(N4)+ float(N5) + float(N6) + float(N8)+float(N9)
    #Fixed Finance
        TotalM= float(M1)+float(M2)+float(M3)+float(M4)
    # Annual Dynamic dues 
        TotalK = float(K1)+float(K2)+float(K3)+float(K4)
        TotalD= TotalN-TotalK
        if "School of Engineering and Applied Sciences (EAS)"in s:
            Major=1
        elif "School of Information Technology and CS (ITCS)"in s:
            Major=2
        elif "School of Business Administration (BA)" in s:
            Major=3
        if Major ==1 : 
            #Need-based Scholarship caculations for Engineering 
            
            if TotalD>= 68400 and TotalM>= 342000 :
                S1=0
            elif TotalD>= 54720 and TotalM>= 273600 :
                S1= 20   
            elif TotalD>= 47880 and TotalM>= 239400 :
                S1= 30
            elif TotalD>= 41040 and TotalM>= 205200 :
                S1= 40
            else :
                S1= 50
                
            #Merit Scholarship caculations For engineering 
            
            if grade>= 95  :
                S2=40
            elif grade>= 90 :
                S2= 35  
            elif grade>= 85  :
                S2= 30
            elif grade>= 80  :
                S2= 25
            else :
                S2= 0 
         
        elif Major == 2 : #Computer Science 
            
            #Need-based Scholarship caculations for Computer Science 
            
            if TotalD>= 56200 and TotalM>= 281000 :
                S1=0
            elif TotalD>= 44960 and TotalM>= 224800 :
                S1= 20   
            elif TotalD>= 39340 and TotalM>= 196700 :
                S1= 30
            elif TotalD>= 33720 and TotalM>= 168600 :
                S1= 40
            else :
                S1= 50
                
            #Merit Scholarship caculations for Computer Science 
            
            if grade>= 95  :
                S2=40
            elif grade>= 90 :
                S2= 35  
            elif grade>= 85  :
                S2= 30
            elif grade>= 80  :
                S2= 25
            else:
                S2= 0
                
            
            
        elif Major == 3 : #Business
            
            #Need-based Scholarship caculations for Business
            if TotalD>= 51600 and TotalM>= 258000 :
                S1=0
            elif TotalD>= 41280 and TotalM>= 206400 :
                S1= 20   
            elif TotalD>= 36120 and TotalM>= 180600 :
                S1= 30
            elif TotalD>= 30960 and TotalM>= 154800 :
                S1= 40
            else :
                S1= 50
                
            #Merit Scholarship caculations for Business 
            
            if grade>= 95  :
                S2=40
            elif grade>= 90 :
                S2= 35  
            elif grade>= 85  :
                S2= 30
            elif grade>= 80  :
                S2= 25
            else :
                S2= 0
        Total_scholarship = S1+S2
        results()
def bchange (*args):
    if e2.get() ==1 and e3.get()==1 and e4.get()==1 and (e12.get()==1 or e12.get()==2)  :        
        cv2.itemconfig(background, image = i23412)
    elif e2.get() ==1 and e3.get()==1 and e4.get()==1 :
        cv2.itemconfig(background, image = i234)
    elif e2.get() ==1 and e3.get()==1 and (e12.get()==1 or e12.get()==2) :
        cv2.itemconfig(background, image = i2312)        
    elif e2.get() ==1 and e3.get()==1:
        cv2.itemconfig(background, image = i23)
    elif e2.get() ==1 and e4.get()==1 and (e12.get()==1 or e12.get()==2):
        cv2.itemconfig(background, image = i2412)        
    elif e2.get() ==1 and e4.get()==1 :
        cv2.itemconfig(background, image = i24)
    elif e2.get() ==1 and (e12.get()==1 or e12.get()==2):
        cv2.itemconfig(background, image = i212) 
    elif e2.get() ==1 :
        cv2.itemconfig(background, image = i2) 
    elif (e12.get() ==1 or e12.get()==2) and e3.get()==1 and e4.get()==1 :
        cv2.itemconfig(background, image = i3412)
    elif e3.get()==1 and e4.get()==1 :
        cv2.itemconfig(background, image = i34)
    elif e3.get()==1 and (e12.get()==1 or e12.get()==2) :
        cv2.itemconfig(background, image = i312) 
    elif e3.get()==1 :
        cv2.itemconfig(background, image = i3)
    elif e4.get() ==1 and (e12.get()==1 or e12.get()==2):
        cv2.itemconfig(background, image = i412) 
    elif e4.get()==1:
        cv2.itemconfig(background, image = i4)
    elif (e12.get()==1 or e12.get()==2):
        cv2.itemconfig(background, image = i12)        
        
def radio2 (*args):
    global e22
    if e2.get() ==1 :
        r1.destroy()
        r2.destroy() 
        e22 = IntVar()
        e22 = Entry(financial_screen,width=9,borderwidth=0)
        e22.place(x=346,y=236)      
        bchange()
        
        
def radio3 (*args):
    global e33
    if e3.get() ==1 :
        r3.destroy()
        r4.destroy()
        e33 = IntVar()
        e33 = Entry(financial_screen,width=9,borderwidth=0)
        e33.place(x=346,y=279)
        bchange()
 
        
        
def radio4 (*args):
    global e44
    if e4.get() ==1 :
        r5.destroy()
        r6.destroy()
        e44 = IntVar()
        e44 = Entry(financial_screen,width=9,borderwidth=0)
        e44.place(x=346,y=371)  
        bchange()
        
        
def radio12 (*args):
    global M3,K1,e1212
    if e12.get() ==1 or e12.get() ==2:
        r7.destroy()
        r8.destroy()
        e1212 = IntVar()
        e1212 = Entry(financial_screen,width=9,borderwidth=0)
        e1212.place(x=760,y=282)      
        bchange()


def  financialscreen (v1,v5,v6,v8,v9,v10,v11,v13,v14,v15,v16,v17,v18):
    global e1,e2,e3,e4,e5,e6,e8,e9,e10,e11,e12,e13,e14,e15,e16,e17,e18,r1,r2,r3,r4,r5,r6,r7,r8,financial_screen,cv2,background,defaultimg,defaultimg3,defaultimg2,i12,i2,i212,i23,i2312,i234,i23412,i24,i2412,i3412,i34,i312,i3,i412,i4,submit,labelback2,studentnamelabel,studentname,studentidlabel,studentnamelabel,studentid
    financial_screen = Tk()
    financial_screen.title("Financial Calculator")
    financial_screen.geometry("922x748")
    financial_screen.grid_columnconfigure(1,weight=1)
    financial_screen.resizable(0, 0)
    center(financial_screen)  
    
    cv2 = tk.Canvas(financial_screen, width=932, height=748)
    cv2.pack(fill='both')  
    defaultimg=ImageTk.PhotoImage(Image.open(mypath+"\\Images\\Financial.jpg" ))
    defaultimg2=ImageTk.PhotoImage(Image.open(mypath+"\\Images\\Financial2.jpg" ))
    defaultimg3=ImageTk.PhotoImage(Image.open(mypath+"\\Images\\Financial3.jpg" ))
    i23412=ImageTk.PhotoImage(Image.open(mypath+"\\Images\\2,3,4,12.jpg"))
    i234=ImageTk.PhotoImage(Image.open(mypath+"\\Images\\2,3,4.jpg"))
    i2312=ImageTk.PhotoImage(Image.open(mypath+"\\Images\\2,3,12.jpg"))
    i23=ImageTk.PhotoImage(Image.open(mypath+"\\Images\\2,3.jpg"))
    i2412=ImageTk.PhotoImage(Image.open(mypath+"\\Images\\2,4,12.jpg"))
    i24=ImageTk.PhotoImage(Image.open(mypath+"\\Images\\2,4.jpg"))
    i212=ImageTk.PhotoImage(Image.open(mypath+"\\Images\\2,12.jpg"))
    i2=ImageTk.PhotoImage(Image.open(mypath+"\\Images\\2.jpg"))
    i3412=ImageTk.PhotoImage(Image.open(mypath+"\\Images\\3,4,12.jpg"))
    i34=ImageTk.PhotoImage(Image.open(mypath+"\\Images\\3,4.jpg"))
    i312=ImageTk.PhotoImage(Image.open(mypath+"\\Images\\3,12.jpg"))
    i3=ImageTk.PhotoImage(Image.open(mypath+"\\Images\\3.jpg"))
    i412=ImageTk.PhotoImage(Image.open(mypath+"\\Images\\4,12.jpg"))
    i4=ImageTk.PhotoImage(Image.open(mypath+"\\Images\\4.jpg"))
    i12=ImageTk.PhotoImage(Image.open(mypath+"\\Images\\12.jpg"))
    background=cv2.create_image(460, 373, image=defaultimg, anchor='center') 
    
    
    
    #-------------- Entries
    e1 = Entry(financial_screen,width=9,borderwidth=0)
    e1.place(x=346,y=193)
    e1.delete(0,END)
    e1.insert(0,v1)    
    e2 = IntVar()
    r1=Radiobutton(financial_screen, text="Yes",padx = 5, variable=e2, value=1 ,background="white",activebackground="white",width=15)
    r1.place(x=280,y=231)
    r2=Radiobutton(financial_screen, text="No",padx = 5, variable=e2, value=2,background="white",activebackground="white")
    r2.place(x=380,y=231)
    
    e3 = IntVar()
    r3=Radiobutton(financial_screen, text="Yes",padx = 5, variable=e3, value=1 ,background="white",activebackground="white",width=15)
    r3.place(x=280,y=277)
    r4=Radiobutton(financial_screen, text="No",padx = 5, variable=e3, value=2,background="white",activebackground="white")
    r4.place(x=380,y=277)
    
    e4 = IntVar()
    r5=Radiobutton(financial_screen, text="Yes",padx = 5, variable=e4, value=1 ,background="white",activebackground="white",width=15)
    r5.place(x=280,y=365)
    r6=Radiobutton(financial_screen, text="No",padx = 5, variable=e4, value=2,background="white",activebackground="white")
    r6.place(x=380,y=365)
    e5 = Entry(financial_screen,width=9,borderwidth=0)
    e5.place(x=346,y=468)
    e5.delete(0,END)
    e5.insert(0,v5)    
    e6 = Entry(financial_screen,width=9,borderwidth=0)
    e6.place(x=345,y=526)
    e6.delete(0,END)
    e6.insert(0,v6)    
    e8 = Entry(financial_screen,width=9,borderwidth=0)
    e8.place(x=346,y=573)
    e8.delete(0,END)
    e8.insert(0,v8)    
    e9 = Entry(financial_screen,width=9,borderwidth=0)
    e9.place(x=346,y=620)
    e9.delete(0,END)
    e9.insert(0,v9)    
    e10 = Entry(financial_screen,width=9,borderwidth=0)
    e10.place(x=762,y=193)
    e10.delete(0,END)
    e10.insert(0,v10)    
    e11 = Entry(financial_screen,width=9,borderwidth=0)
    e11.place(x=762,y=236)
    e11.delete(0,END)
    e11.insert(0,v11)    
    e12 = IntVar()
    r7=Radiobutton(financial_screen, text="Rented",padx = 5, variable=e12, value=1 ,background="white",activebackground="white",width=15)
    r7.place(x=698,y=277)
    r8=Radiobutton(financial_screen, text="Owned",padx = 5, variable=e12, value=2,background="white",activebackground="white")
    r8.place(x=798,y=277)
    e13 = Entry(financial_screen,width=9,borderwidth=0)
    e13.place(x=762,y=329)
    e13.delete(0,END)
    e13.insert(0,v13)    
    e14 = Entry(financial_screen,width=9,borderwidth=0)
    e14.place(x=762,y=420)
    e14.delete(0,END)
    e14.insert(0,v14)    
    e15 = Entry(financial_screen,width=9,borderwidth=0)
    e15.place(x=762,y=470)
    e15.delete(0,END)
    e15.insert(0,v15)    
    e16 = Entry(financial_screen,width=9,borderwidth=0)
    e16.place(x=762,y=527)
    e16.delete(0,END)
    e16.insert(0,v16)    
    e17 = Entry(financial_screen,width=9,borderwidth=0)
    e17.place(x=762,y=572)
    e17.delete(0,END)
    e17.insert(0,v17)    
    e18 = Entry(financial_screen,width=7,borderwidth=0)
    e18.place(x=762,y=620)
    e18.delete(0,END)
    e18.insert(0,v18)    
    e2.trace('w', radio2)
    e3.trace('w', radio3)
    e4.trace('w', radio4)
    e12.trace('w', radio12)
     
    submit=Button(financial_screen, font=("bold", 13), text='Calculate',width=20,height=1,bg='#002145',fg='white')
    submit.place(x=355,y=680)
    submit.bind("<Button-1>",getdata)
    financial_screen.bind("<Return>",getdata)    
    labelback2 = Label(financial_screen, text="Back",width=8,font=("bold", 13),bg="#002145",fg="white")
    labelback2.bind("<Button-1>", back2)
    labelback2.place(x=424,y=10)
    
    studentnamelabel=Label(text="Student Name :", width="11", height="1", font=("Lato", 10),background="white",activebackground="white")
    studentnamelabel.place(x=370,y=65)
     
    studentname=Label(text=stuname, width="20", height="1", font=("Lato", 10),background="white",activebackground="white")
    studentname.place(x=464,y=65)
    studentname.config(anchor='w')
    studentidlabel=Label(text="Student ID :", width="10", height="1", font=("Lato", 10),background="white",activebackground="white")
    studentidlabel.place(x=384,y=95)
        
    studentid=Label(text=stuID, width="15", height="1", font=("Lato", 10),background="white",activebackground="white")
    studentid.place(x=464,y=95)
    studentid.config(anchor='w')
    

    financial_screen.mainloop()        
    
    
    
#---------------------------------------------------------------------------------------------------------------------------------   

def career_calculator(id):
    wb = openpyxl.load_workbook(mypath+"\\Analysis\\Career_Recommendation.xlsx")
    Students = wb["Students"]
    Careers= wb["Careers"]
    val=0
    valjob=0
    Job_description=[]
    Students_abilities=[]
    percentage1=[]
    for j in range(2,12):
        for i in range(1,15):
            V=Careers.cell(row=j, column=i).value
            valjob+=V
        Job_description.append(valjob)
        valjob=0
    for j in range(2,21):
        for i in range(1,15):
            V=Students.cell(row=j, column=i).value
            val+=V
        Students_abilities.append(val)
        val=0
    for i in Students_abilities:
        for j in Job_description:
            percentage1.append(abs(j-i))
    for i,j in zip(range(37,48), range(2,12)):
        Students.cell(row=1, column=i).value= Careers.cell(row=j, column=37).value
    val2=0
    List2=[]
    for s in range(2,21):
        for i in range(2,12):
            for j in range(15,37):
                k= Careers.cell(row=i, column=j).value
                if k == Students.cell(row=s, column=j).value:
                    val2+=1
            if Careers.cell(row=i, column=34).value == Students.cell(row=s, column=34).value:
                val2+=10
            if Careers.cell(row=i, column=33).value == Students.cell(row=s, column=33).value:
                val2+=10
            if Careers.cell(row=i, column=32).value == Students.cell(row=s, column=32).value:
                val2+=10
            List2.append(-val2)
            val2=0
    a=0
    b=11
    for j in range(2,21):
        for i,q,o in zip(percentage1[a:b], range(37,47), List2):
            Students.cell(row=j, column=q).value= i+o
        a+=10
        b+=10
    LO=[]
    for i in range(2,21):
        for j in range(37,47):
            LO.append(Students.cell(row=i, column=j).value)
            LO2=LO[ : :1]
        Students.cell(row=i, column=47).value=Students.cell(row=1, column=LO2.index(min(LO))+37).value
        LO.remove(min(LO))
        Students.cell(row=i, column=48).value=Students.cell(row=1, column=LO2.index(min(LO))+37).value
        LO.remove(min(LO))
        Students.cell(row=i, column=49).value=Students.cell(row=1, column=LO2.index(min(LO))+37).value
        LO=[]
        LO2=[]
    #wb.save('example.xlsx')
    def ID(ID):
        for i in range(2,21):
            if ID == Students.cell(row=i, column=51).value:
                return [Students.cell(row=i, column=47).value, Students.cell(row=i, column=48).value, Students.cell(row=i, column=49).value]
                       
    return (ID(id))
def career_results():
    global tex13,career_screen,progress2,cv3,background3,x
    Home_screen.destroy()
    career_screen = Tk()
    career_screen.title("Career Recommendation")
    career_screen.geometry("922x748")
    career_screen.grid_columnconfigure(1,weight=1)
    career_screen.resizable(0, 0)
    center(career_screen)  

    cv3 = tk.Canvas(career_screen, width=932, height=748)
    cv3.pack(fill='both')
    defaultimg7=ImageTk.PhotoImage(Image.open(mypath+"\\Images\\Careerb.jpg" ))
    ref = Label(image=defaultimg7)
    ref.image = defaultimg7 # Just keeping a reference!    
    background3=cv3.create_image(460, 373, image=defaultimg7, anchor='center')             
        
    progress2 = ttk.Progressbar(career_screen, orient = HORIZONTAL, length = 450)
    progress2.place(x=236,y=254)
    progress2.config(mode = 'determinate',maximum=70, value = 1)
    progress2.step(1)
    pygame.mixer.Sound(mypath+"\\Sound\\analysing.wav").play() 
    progress2.start()
    x= career_calculator(int(stuID))
    tex13=Label(text="Collecting student Information", width="24", height="2", font=("Lato", 12),background="white",activebackground="white")
    tex13.place(x=360,y=280)    
    progress2.after(3400, changetext2)
    progress2.after(6800, afterprogress2)
         
def changetext2():
    global tex14
    tex13.destroy()
    tex14=Label(text="Analysing grades and activities ", width="30", height="2", font=("Lato", 12),background="white",activebackground="white")
    tex14.place(x=330,y=280)   
def afterprogress2():
    progress2.stop()
    progress2.destroy()
    tex14.destroy()
    pygame.mixer.Sound(mypath+"\\Sound\\analysiscomp.wav").play()
    labelback3 = Label(career_screen, text="Back",width=8,font=("bold", 13),bg="#002145",fg="white")
    labelback3.bind("<Button-1>", back3)
    labelback3.place(x=424,y=10)    
    if x!=None:
        defult4=ImageTk.PhotoImage(Image.open(mypath+"\\Images\\Careerb1.jpg")) 
        ref = Label(image=defult4)
        ref.image = defult4 # Just keeping a reference!    
        cv3.itemconfig(background3, image = defult4)
    
        tex10=Label(career_screen,text=x[0], width="16", height="2", font=("Lato", 14),background="#f6f6f6",activebackground="white")
        tex10.place(x=250,y=258)
        tex10.config(anchor='w')
        tex11=Label(career_screen,text=x[1], width="16", height="1", font=("Lato", 14),background="#f6f6f6",activebackground="white")
        tex11.place(x=240,y=430)
        tex11.config(anchor='w')
        tex12=Label(career_screen,text=x[2], width="16", height="1", font=("Lato", 14),background="#f6f6f6",activebackground="white")
        tex12.place(x=250,y=585)
        tex12.config(anchor='w')
    
    else:
        defult5=ImageTk.PhotoImage(Image.open(mypath+"\\Images\\Careerb2.jpg"))
        ref = Label(image=defult5)
        ref.image = defult5 # Just keeping a reference!    
        cv3.itemconfig(background3, image = defult5)
#---------------------------------------------------------------------------------------------------------
def sentianly():
        
  # opening the spreedsheet to read the data   
    wb = openpyxl.load_workbook(file)
    sheetname = str(wb.get_sheet_names()[0])
    sheet = wb.get_sheet_by_name(sheetname)

    # defining our counters
    pos = 0
    neg = 0
    normal = 0
    for i in range(2, 500):

        # looping on a range of cells where the feedbacks are located 
            value = sheet.cell(row=i, column=2).value
            if value != None:
                vs = analyzer.polarity_scores(value)

                #conunting the number of positive, negative, and normal feedbacks
                if vs["compound"] >= 0.05:
                    
                    pos += 1
                elif vs["compound"] > -0.05 and vs["compound"] < 0.05:
                    
                    normal += 1
                elif vs["compound"] <= -0.05:
                    
                    neg +=1
                    
    # preparing the data for graph plotting
    sum = pos + neg + normal
    
    pos_percent = int(pos / sum * 100)
    neg_percent = int(neg / sum * 100)
    normal_percent = int(normal / sum * 100)

    figsize = (4.4/3.3, 1)
    plt.figure(figsize=figsize,dpi=100)    
    # Pie chart, where the slices will be ordered and plotted counter-clockwise:
    labels = 'Positive', 'Negative', 'Neutral'
    sizes = [pos_percent, neg_percent, normal_percent]
    explode = (0.11, 0, 0)  # only "explode" the 2nd slice (i.e. 'Hogs')
    
    fig1, ax1 = plt.subplots()
    ax1.pie(sizes, explode=explode, labels=labels, autopct='%1.1f%%',
            shadow=True, startangle=90)
    ax1.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.

    #The Outputs for the GUI

    print(pos ,neg ,normal) # The Number of Positive, Negative, and Neutral Feedbacks.
    print(sum) # The Total Number of feedbacks analyzed
    plt.savefig(mypath+"\\Images\\PieChart.png",type="png",dpi=75)   

def openanalyser():
    global progress3,tex15,senti_screen
    click1.play()
    Home_screen.destroy()
    
    global senti_screen,cv4,background4,sentimage1,sentimage2
    senti_screen = Tk()
    senti_screen.title("Sentimental Analyzer")
    senti_screen.geometry("922x748")
    senti_screen.grid_columnconfigure(1,weight=1)
    senti_screen.resizable(0, 0)
    center(senti_screen)    
    cv4 = tk.Canvas(senti_screen, width=932, height=748)
    cv4.pack(fill='both')
    sentimage1=ImageTk.PhotoImage(Image.open(mypath+"\\Images\\sent1.jpg"))
    sentimage2= ImageTk.PhotoImage(Image.open(mypath+"\\Images\\sent2.jpg"))
    ref = Label(image=sentimage1)
    ref.image = sentimage1 # Just keeping a reference!    
    background4=cv4.create_image(460, 373, image=sentimage1, anchor='center')
    
    progress3 = ttk.Progressbar(senti_screen, orient = HORIZONTAL, length = 450)
    progress3.place(x=236,y=254)
    progress3.config(mode = 'determinate',maximum=70, value = 1)
    progress3.step(1)
    progress3.start()
    pygame.mixer.Sound(mypath+"\\Sound\\analysing.wav").play()
    tex15=Label(text="Bulding Machine Learning Model", width="30", height="2", font=("Lato", 12),background="white",activebackground="white")
    tex15.place(x=335,y=280)    
    progress3.after(3310,sentimentalf)

def sentimentalf ():
    global analyzer
    # defining the sentiment analyzer -------
    analyzer = SentimentIntensityAnalyzer()
    
    pos_count = 0
    pos_correct = 0
    
    # training on the positive  datasets
    with open(mypath+"\\Analysis\\positive_dataset.txt","r") as f:
        for line in f.read().split('\n'):
            vs = analyzer.polarity_scores(line)
            if not vs['neg'] > 0.1:
                if vs['pos']-vs['neg'] > 0:
                    pos_correct += 1
                pos_count +=1
    
    
    neg_count = 0
    neg_correct = 0
    
    # training on the negative  datasets
    
    with open(mypath+"\\Analysis\\negative_dataset.txt","r") as f:
        for line in f.read().split('\n'):
            vs = analyzer.polarity_scores(line)
            if not vs['pos'] > 0.1:
                if vs['pos']-vs['neg'] <= 0:
                    neg_correct += 1
                neg_count +=1    
    afterprogress3()     


    senti_screen.mainloop()

def afterprogress3():
    global file
    progress3.stop()
    progress3.destroy()
    tex15.destroy()
    # programming the button to call the attach action
    messagebox.showinfo("Feedback File", "Please choose the feedback file you want to analyze")
    file = filedialog.askopenfilename()
    sentianly()
    cv4.itemconfig(background4, image = sentimage2 )
    labelback4 = Label(senti_screen, text="Back",width=8,font=("bold", 13),bg="#002145",fg="white")
    labelback4.bind("<Button-1>", back4)
    labelback4.place(x=424,y=10)    
    ImageNew3 = ImageTk.PhotoImage(Image.open(mypath+"\\Images\\PieChart.png"))
    mat = Canvas(senti_screen, width = 440, height = 330,highlightthickness=0)
    mat.place(x=246,y=217)
    mat.create_image(226,170,anchor='center',image=ImageNew3)
    mat.image = ImageNew3
    pygame.mixer.Sound(mypath+"\\Sound\\analysiscomp.wav").play()
def getfinancial():
    click1.play()
    Home_screen.destroy()
    financialscreen("","","","","","","","","","","","","")
def logout(Event):
    click2.play()
    Home_screen.destroy()
    main_account_screen() 
def Hover(event,me,newpath):
    pygame.mixer.Sound(mypath+"\\Sound\\hover.wav").play()
    me.delete(me.find_all())
    me_ImageNew = ImageTk.PhotoImage(Image.open(newpath))
    me.create_image(0,1,anchor=NW,image=me_ImageNew)
    me.image = me_ImageNew
    
def Leave(event,me,newpath):
    me.delete(me.find_all())
    me_ImageNew = ImageTk.PhotoImage(Image.open(newpath))
    me.create_image(0,1,anchor=NW,image=me_ImageNew)
    me.image = me_ImageNew    
def botclick(event):
    click1.play()
   
def Release (event,me, palce):
    me.after(110, Place)     
def Home_screenf():
    global Home_screen,stuID,stuname
    Home_screen = Tk()
    Home_screen.title("NU continuum")
    Home_screen.geometry("861x678")  
    Home_screen.resizable(0, 0)

    cv = tk.Canvas(Home_screen, width=273, height=685)
    cv.pack(fill='both')
    img_file = mypath+"\\Images\\Main.jpg"
    pil_image = Image.open(img_file)   
    tk_image = ImageTk.PhotoImage(pil_image)
    cv.create_image(430, 333, image=tk_image, anchor='center')

    logoutlabel=Label(text="Logout", width="10", height="1", font=("Lato", 10),background="#002145",fg="white")
    logoutlabel.place(x=35,y=27)
    logoutlabel.bind("<Button-1>",logout)
    pygame.mixer.Sound(mypath+"\\Sound\\how can we.wav").play()
    k=[]
    for i in s:
        x=i.split(" ")
        k.append(x[0])
        k.append(x[1])
    stuIDindex= k.index("ID:")+1
    stuID=k[stuIDindex]
    
    for i in s:
        x=i.split(" ")
        if "Name:" in x:
            c=x[1: ]
            stuname=" ".join(c)    
    boarder = ImageTk.PhotoImage(Image.open(mypath+"\\Images\\Career1.png"))
    career = Canvas(Home_screen, width = 315, height = 65,bg="white",highlightthickness=0)
    career.place(x=272,y=210)
    career.create_image(0,0,anchor=NW,image=boarder)
    career.image = boarder
    Label_career=Label(text="Career Recommendation", width="20", height="2", font=("Lato", 12),background="white",activebackground="white")
    Label_career.place(x=343,y=225)
    career,Label_career.bind("<Button-1>",botclick)
    career,Label_career.bind("<Enter>",lambda event: Hover(event,career,mypath+"\\Images\\Career2.png"))
    career,Label_career.bind("<Leave>",lambda event: Leave(event,career, mypath+"\\Images\\Career1.png"))  
    career,Label_career.bind("<ButtonRelease-1>",lambda event: Release(event,career, career_results()))  
#------------------------------------- bot2

    financial = Canvas(Home_screen, width = 315, height = 65,bg="white",highlightthickness=0)
    financial.place(x=272,y=330)
    financial.create_image(0,0,anchor=NW,image=boarder)
    financial.image = boarder
    Label_financial=Label(text="Financial Aid Calculator", width="20", height="2", font=("Lato", 12),background="white",activebackground="white")
    Label_financial.place(x=343,y=345)
    financial,Label_financial.bind("<Button-1>",botclick)
    financial,Label_financial.bind("<Enter>",lambda event: Hover(event,financial,mypath+"\\Images\\Career2.png"))
    financial,Label_financial.bind("<Leave>",lambda event: Leave(event,financial, mypath+"\\Images\\Career1.png"))  
    financial,Label_financial.bind("<ButtonRelease-1>",lambda event: Release(event,financial, getfinancial()))  
#---------------------------------------bot3
    bot3= Canvas(Home_screen, width = 315, height = 65,bg="white",highlightthickness=0)
    bot3.place(x=272,y=450)
    bot3.create_image(0,0,anchor=NW,image=boarder)
    bot3.image = boarder
    Label_bot3=Label(text="What People think of us", width="20", height="2", font=("Lato", 12),background="white",activebackground="white")
    Label_bot3.place(x=343,y=465)
    bot3,Label_bot3.bind("<Button-1>",botclick)
    bot3,Label_bot3.bind("<Enter>",lambda event: Hover(event,bot3,mypath+"\\Images\\Career2.png"))
    bot3,Label_bot3.bind("<Leave>",lambda event: Leave(event,bot3, mypath+"\\Images\\Career1.png"))  
    bot3,Label_bot3.bind("<ButtonRelease-1>",lambda event: Release(event,bot3, openanalyser()))  

    center(Home_screen)     
    Home_screen.mainloop()    

def back2(Event):
    click2.play()
    financial_screen.destroy()
    Home_screenf()
def back3(Event):
    click2.play()
    career_screen.destroy()
    Home_screenf()
def back4(Event):
    click2.play()
    senti_screen.destroy()
    Home_screenf()    
def password_not_recognised():
    global password_not_recog_screen
    password_not_recog_screen = Toplevel(main_screen)
    password_not_recog_screen.title("Error")
    password_not_recog_screen.geometry("210x60")
    Label(password_not_recog_screen, text="Invalid Password ").pack()
    Button(password_not_recog_screen, text="OK", command=delete_password_not_recognised).pack()
    center(password_not_recog_screen)    
 
 

 
def delete_password_not_recognised():
    password_not_recog_screen.destroy()
 
 
 
def Release (event,me, palce):
    me.after(110, Place)

    
def Click(event,me,newpath): 
    click1.play()
    me.delete(me.find_all())
    me_ImageNew = ImageTk.PhotoImage(Image.open(newpath))
    me.create_image(0,1,anchor=NW,image=me_ImageNew)
    me.image = me_ImageNew
    
def main_account_screen():
    global main_screen, B1,B2,Label1,Label2,NewButton,NewButton2
    main_screen = Tk()
    main_screen.geometry("485x450")
    main_screen.title("Interface")
    main_screen.resizable(0, 0)
    cv = tk.Canvas(main_screen, width=960, height=600)
    cv.pack(fill='both')
    img_file = mypath+"\\Images\\Untitled-2.jpg"
    pil_image = Image.open(img_file)  
    tk_image = ImageTk.PhotoImage(pil_image)
    cv.create_image(245, 224, image=tk_image, anchor='center')    
    center(main_screen)    
    
    ImageNew = ImageTk.PhotoImage(Image.open(mypath+"\\Images\\Login1.png"))
    NewButton = Canvas(main_screen, width = 274, height = 78,bg="white",highlightthickness=0)
    NewButton.place(x=105,y=265)
    NewButton.create_image(0,0,anchor=NW,image=ImageNew)
    NewButton.image = ImageNew      
    NewButton.bind("<Button-1>",lambda event: Click(event,NewButton,mypath+"\\Images\\Login2.png"))
    NewButton.bind("<ButtonRelease-1>",lambda event:Release (event,NewButton2, login()))     
    
    ImageNew2 = ImageTk.PhotoImage(Image.open(mypath+"\\Images\\Signin1.png"))
    NewButton2 = Canvas(main_screen, width = 215, height = 55,bg="white",highlightthickness=0)
    NewButton2.place(x=130,y=370)
    NewButton2.create_image(0,-12,anchor=NW,image=ImageNew2)
    NewButton2.image = ImageNew2
    NewButton2.bind("<Button-1>",lambda event: Click(event,NewButton2,mypath+"\\Images\\Signin2.png"))
    NewButton2.bind("<ButtonRelease-1>",lambda event: Release(event,NewButton2, register()))       
    
    Label1=Label(text="Select Your Choice", width="30", height="2", font=("Calibri", 12),bg="white")
    Label1.place(x=124,y=228)
    Label2= Label(text="or", width="5", height="1", font=("Calibri", 12),bg="white")
    Label2.place(x=220,y=340)
    pygame.mixer.Sound(mypath+"\\Sound\\Welcome.wav").play()  
    main_screen.mainloop()
 
main_account_screen()    
 
 
 
